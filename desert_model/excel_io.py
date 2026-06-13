"""Excel output helpers for Result.xlsx."""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Dict, Mapping

from .models import SolutionTrace


RESULT_TEMPLATE = Path("Result.xlsx")


def write_result_workbook(
    template_path: Path,
    output_path: Path,
    traces: Mapping[int, SolutionTrace],
) -> None:
    """Fill the official Result.xlsx template for levels 1 and 2."""

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write xlsx files; use the bundled Python runtime") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    layout = {
        1: {"row0": 4, "node": 2, "cash": 3, "water": 4, "food": 5},
        2: {"row0": 4, "node": 8, "cash": 9, "water": 10, "food": 11},
    }
    for level_id, trace in traces.items():
        if level_id not in layout or not trace.feasible:
            continue
        columns = layout[level_id]
        for step in trace.steps:
            row = columns["row0"] + step.day
            sheet.cell(row=row, column=columns["node"]).value = step.state.node
            sheet.cell(row=row, column=columns["cash"]).value = step.state.cash
            sheet.cell(row=row, column=columns["water"]).value = step.state.water
            sheet.cell(row=row, column=columns["food"]).value = step.state.food
        final_day = trace.steps[-1].day if trace.steps else -1
        for day in range(final_day + 1, 31):
            row = columns["row0"] + day
            for col in (columns["node"], columns["cash"], columns["water"], columns["food"]):
                sheet.cell(row=row, column=col).value = None
    workbook.save(output_path)


def write_solve_status(path: Path, traces: Mapping[int, SolutionTrace]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        str(level_id): {
            "feasible": trace.feasible,
            "status": trace.status,
            "message": trace.message,
            "objective_value": trace.objective_value if math.isfinite(trace.objective_value) else None,
            "final_state": trace.final_state().__dict__ if trace.final_state() else None,
            "metadata": trace.metadata,
        }
        for level_id, trace in sorted(traces.items())
    }
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, allow_nan=False)
        handle.write("\n")
